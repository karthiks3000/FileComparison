#The Code will compare the files and folders in Scrounger NY with Scrounger AWS

# Using Python3
import time
import threading
from queue import Queue
from tkinter import Tk,filedialog  ## notice lowercase 't' in tkinter here

import os
import scandir
import hashlib
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,colors
from tqdm import tqdm


class FileData:
    SourceFileName = ''
    DestinationFileName = ''
    SourceMD5 = ''
    DestinationMD5 =''
    ComparisonResult = ''

    # function to calculate the checksum value of the provided file path
    def md5Checksum(self, file_name):
        with open(file_name, 'rb') as fh:
            m = hashlib.md5()
            while True:
                data = fh.read(8192)
                if not data:
                    break
                m.update(data)
            return m.hexdigest()
    
    
# code starts here
class FileComparison:
    
    root = Tk()
    root.withdraw()
    folder_size_source = 0
    folder_size_destination = 0
    lock = threading.Lock()

    def CheckFile(self,i,path,file):
        fileObject = FileData()                 

        fileObject.SourceFileName = os.path.join(path, file)
        fileObject.DestinationFileName = os.path.join(self.destinationFolder, os.path.relpath(path,self.sourceFolder))
        fileObject.DestinationFileName = os.path.join(fileObject.DestinationFileName,file)

        with self.lock:
            self.folder_size_source += os.path.getsize(fileObject.SourceFileName)

        if(os.path.isfile(fileObject.DestinationFileName)):                
            with self.lock:
                self.folder_size_destination += os.path.getsize(fileObject.DestinationFileName)    

            fileObject.SourceMD5 =  fileObject.md5Checksum(fileObject.SourceFileName)
            fileObject.DestinationMD5 =  fileObject.md5Checksum(fileObject.DestinationFileName)  
            
            #Compare the checksum of Source and Files
            if(fileObject.SourceMD5 != fileObject.DestinationMD5):
                # sheet1.cell(i,5,'Checksum does not match')
                fileObject.ComparisonResult = 'Checksum does not match'
            else:
                # sheet1.cell(i,5,'Checksum matched')
                fileObject.ComparisonResult = 'Checksum matched'
                                
        else:
            # sheet1.cell(i,5,destination_filename + ' does not exist')            
            fileObject.ComparisonResult = fileObject.DestinationFileName + ' does not exist'

        self.listOfFiles.insert(i,fileObject)


    def process_queue(self):
        while True:
            i,path,file = self.compress_queue.get()
            self.CheckFile(i,path,file)
            self.compress_queue.task_done()

    def Start(self):
        
        # Selects the Source Folder
        self.sourceFolder = filedialog.askdirectory()
        # Selects the Destination Folder
        self.destinationFolder = filedialog.askdirectory()
        print('Source Path: ' + self.sourceFolder)
        print('')

        print('Destination Path: ' + self.destinationFolder)
        print('')

        self.listOfFiles = []
        i = 0
        # for each file in source, compare with destination   
        start = time.time()    
        
        print('Starting lookup....')       

        self.compress_queue = Queue()

        for i in range(30):
            t = threading.Thread(target=self.process_queue)
            t.daemon = True
            t.start()

        for (path, dirs, files) in tqdm(scandir.walk(self.sourceFolder)):
            for file in files:        
                self.compress_queue.put((i,path,file))
                i = i + 1
                
        self.compress_queue.join()                
        # logic to print data to excel
        print('Hang tight...Printing results to excel....')

        self.book = Workbook()

        f = Font(name='Calibri',
        size=11,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=colors.RED)  

        sheet1 = self.book.create_sheet('Results')

        sheet1.cell(1,1,'Source Folder: ')
        sheet1.cell(1,1).font= f
        sheet1.cell(1,2, self.sourceFolder)

        sheet1.cell(2,1,'Destination Folder:')
        sheet1.cell(2,1).font= f
        sheet1.cell(2,2, self.destinationFolder)

        sheet1.cell(5,1,'File Comparison Results:')
        sheet1.cell(5,1).font= f

        sheet1.cell(7,1,'Source File Path')
        sheet1.cell(7,1).font= f

        sheet1.cell(7,2,'Source File MD5')
        sheet1.cell(7,2).font= f

        sheet1.cell(7,3,'Destination File Path')
        sheet1.cell(7,3).font= f

        sheet1.cell(7,4,'Destination File MD5')
        sheet1.cell(7,4).font= f

        sheet1.cell(7,5,'Result')
        sheet1.cell(7,5).font= f

        sheet1.column_dimensions["A"].width = 50
        sheet1.column_dimensions["B"].width = 35
        sheet1.column_dimensions["C"].width = 50
        sheet1.column_dimensions["D"].width = 35
        sheet1.column_dimensions["E"].width = 50

        i = 8

        # add data from list

        for file in self.listOfFiles:
            sheet1.cell(i,1,file.SourceFileName)
            sheet1.cell(i,2,file.SourceMD5)
            sheet1.cell(i,3,file.DestinationFileName)
            sheet1.cell(i,4,file.DestinationMD5)
            sheet1.cell(i,5,file.ComparisonResult)
            i = i + 1
                
        i=i+3
        sheet1.cell(i,1,'Size of files found in source')
        sheet1.cell(i,1).font= f

        sheet1.cell(i,2,'%0.1f MB' % (self.folder_size_source/(1024*1024.0)))

        i=i+1
        sheet1.cell(i,1,'Size of files found in destination')
        sheet1.cell(i,1).font= f

        sheet1.cell(i,2,'%0.1f MB' % (self.folder_size_destination/(1024*1024.0)))        
        
        print('')
        print("Source Folder Size: %0.1f MB" % (self.folder_size_source/(1024*1024.0)))
        print("Destination Folder Size = %0.1f MB" % (self.folder_size_destination/(1024*1024.0)))

        self.book.save('Result.xlsx')
        print("Execution time = {0:.5f}".format(time.time() - start))
        


f = FileComparison()
f.Start()
